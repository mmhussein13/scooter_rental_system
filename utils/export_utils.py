"""Utility functions for exporting data to Excel"""
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def export_to_excel(data, columns, filename, title=None, sheet_name="Sheet1"):
    """
    Export data to an Excel file
    
    Args:
        data: List of dictionaries containing the data to export
        columns: List of tuples (column_name, display_name)
        filename: Name of the file to export (without extension)
        title: Title to display at the top of the Excel file (optional)
        sheet_name: Name of the sheet (optional)
        
    Returns:
        HttpResponse object with Excel file
    """
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    
    # Define styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font = Font(name='Calibri', bold=True, size=14)
    title_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(vertical='center')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Row counter
    row_num = 1
    
    # Add title if provided
    if title:
        worksheet.merge_cells(f'A{row_num}:{get_column_letter(len(columns))}{row_num}')
        title_cell = worksheet.cell(row=row_num, column=1)
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        row_num += 2  # Leave an empty row after the title
    
    # Add header row
    for col_num, (col_name, display_name) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = display_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        
        # Set column width based on display name length
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = max(15, len(display_name) + 5)
    
    row_num += 1
    
    # Add data rows
    for row_data in data:
        for col_num, (col_name, _) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            
            # Get value from data dictionary, handle nested attributes with dots
            if '.' in col_name:
                parts = col_name.split('.')
                value = row_data
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    elif isinstance(value, dict) and part in value:
                        value = value[part]
                    else:
                        value = None
                        break
            else:
                value = getattr(row_data, col_name) if hasattr(row_data, col_name) else row_data.get(col_name, '')
            
            # Format specific types
            if isinstance(value, datetime.datetime):
                cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(value, datetime.date):
                cell.value = value.strftime('%Y-%m-%d')
            elif value is None:
                cell.value = ''
            else:
                cell.value = value
                
            # Apply styling
            cell.alignment = data_alignment
            cell.border = border
            
        row_num += 1
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response